from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles.fills import PatternFill


class FileSvc:
    def __init__(self):
        pass

    def parse_excel_file(self, filename):
        # Загружаем документ
        book = load_workbook(filename)

        # Определяем рабочий лист
        ws = book.worksheets[0]

        # Определяем необходимую ячейку на листе
        _cell = ws['C2']

        # Задаем стиль для ячейки
        # _cell.font = Font(size=10, underline='single', color='FFFFFF', bold=True, italic=True)
        _cell.value = 200

        # Задаем цвет фона
        # _cell.fill = PatternFill(bgColor="FFC7CE", fill_type="solid")

        # Указываем ширину для колонки
        # ws.column_dimensions["C"].width = 60.0

        # Сохраняем документ
        book.save(filename)

